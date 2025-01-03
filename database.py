import sqlite3
import os
import openpyxl
from datetime import datetime
from collections import defaultdict

DB_NAME = "pallets.db"

def get_db_connection():
    return sqlite3.connect(DB_NAME)

def create_db_if_not_exists():
    """
    Crée la base Pallets / Articles / Movements / Metrics 
    sans notion de threshold, etc.
    """
    first_creation = not os.path.exists(DB_NAME)
    with get_db_connection() as conn:
        c = conn.cursor()

        # Pallets
        c.execute("""
            CREATE TABLE IF NOT EXISTS Pallets (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                bin_name TEXT UNIQUE NOT NULL,
                weight REAL DEFAULT 0,
                image_path TEXT DEFAULT NULL
            )
        """)

        # Articles
        c.execute("""
            CREATE TABLE IF NOT EXISTS Articles (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                bin_id INTEGER NOT NULL,
                code TEXT NOT NULL,
                reference TEXT,
                login TEXT,
                quantity INTEGER DEFAULT 1,
                FOREIGN KEY(bin_id) REFERENCES Pallets(id)
            )
        """)

        # Metrics (pour top 5, etc.)
        c.execute("""
            CREATE TABLE IF NOT EXISTS Metrics (
                id INTEGER PRIMARY KEY,
                articles_in INTEGER DEFAULT 0,
                articles_out INTEGER DEFAULT 0
            )
        """)
        c.execute("SELECT id FROM Metrics WHERE id=1")
        if not c.fetchone():
            c.execute("INSERT INTO Metrics(id, articles_in, articles_out) VALUES(1,0,0)")

        # Movements
        c.execute("""
            CREATE TABLE IF NOT EXISTS Movements (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                article_id INTEGER,
                bin_id INTEGER,
                action TEXT,   -- 'IN' ou 'OUT'
                qty_change INTEGER DEFAULT 0, 
                date_time TEXT,
                FOREIGN KEY(article_id) REFERENCES Articles(id),
                FOREIGN KEY(bin_id) REFERENCES Pallets(id)
            )
        """)

        conn.commit()

def get_or_create_bin(bin_name):
    conn=get_db_connection()
    c=conn.cursor()
    c.execute("SELECT id FROM Pallets WHERE bin_name=?", (bin_name,))
    row=c.fetchone()
    if row:
        bin_id=row[0]
    else:
        c.execute("INSERT INTO Pallets(bin_name) VALUES(?)", (bin_name,))
        bin_id=c.lastrowid
        conn.commit()
    conn.close()
    return bin_id

def get_bin_info(bin_id):
    conn=get_db_connection()
    c=conn.cursor()
    c.execute("SELECT id, bin_name, weight, image_path FROM Pallets WHERE id=?", (bin_id,))
    row=c.fetchone()
    conn.close()
    return row

def get_bin_weight(bin_name):
    conn=get_db_connection()
    c=conn.cursor()
    c.execute("SELECT weight FROM Pallets WHERE bin_name=?", (bin_name,))
    row=c.fetchone()
    conn.close()
    if row:
        return row[0]
    return 0

def update_bin_weight(bin_id, new_weight):
    """
    Met à jour le weight dans Pallets (sans générer de mouvement).
    """
    conn=get_db_connection()
    c=conn.cursor()
    try:
        c.execute("UPDATE Pallets SET weight=? WHERE id=?", (new_weight, bin_id))
        conn.commit()
        return True, ""
    except Exception as e:
        return False, str(e)
    finally:
        conn.close()

def list_articles_in_bin(bin_id):
    """
    Retourne la liste d'articles (id, code, ref, login, quantity).
    """
    conn=get_db_connection()
    c=conn.cursor()
    c.execute("SELECT id, code, reference, login, quantity FROM Articles WHERE bin_id=?", (bin_id,))
    rows=c.fetchall()
    conn.close()
    return rows

def add_article(bin_id, code, reference, login, quantity):
    """
    Ajoute un article => metrics.in++ => Movements(action='IN', qty_change=quantity).
    Si le bin était vide (0 article), on force l'utilisateur (ou on peut forcer le code) 
    à mettre un weight. (Ici, on ne le fait pas automatiquement, 
    juste on peut le signaler dans app.py)
    """
    conn=get_db_connection()
    c=conn.cursor()
    c.execute("""INSERT INTO Articles(bin_id, code, reference, login, quantity)
                 VALUES(?,?,?,?,?)""",
              (bin_id, code, reference, login, quantity))
    art_id=c.lastrowid

    # maj metrics
    c.execute("UPDATE Metrics SET articles_in=articles_in+? WHERE id=1", (quantity,))

    now_str=datetime.now().isoformat()
    c.execute("""INSERT INTO Movements(article_id, bin_id, action, qty_change, date_time)
                 VALUES(?,?,?,?,?)""",
              (art_id, bin_id, 'IN', quantity, now_str))

    conn.commit()
    conn.close()

def remove_article(article_id):
    """
    Supprime un article => on considère tout son quantity comme un 'OUT'.
    S'il n'y a plus d'articles => bin.weight=0
    """
    conn=get_db_connection()
    c=conn.cursor()

    c.execute("SELECT bin_id, quantity FROM Articles WHERE id=?", (article_id,))
    row=c.fetchone()
    if not row:
        conn.close()
        return
    bin_id, old_qty = row[0], row[1]

    # delete l'article
    c.execute("DELETE FROM Articles WHERE id=?", (article_id,))

    # maj metrics => articles_out += old_qty
    c.execute("UPDATE Metrics SET articles_out=articles_out+? WHERE id=1", (old_qty,))

    # Movements => 'OUT' , qty_change= old_qty
    now_str=datetime.now().isoformat()
    c.execute("""INSERT INTO Movements(article_id, bin_id, action, qty_change, date_time)
                 VALUES(?,?,?,?,?)""",
              (article_id, bin_id, 'OUT', old_qty, now_str))

    # verif s'il reste des articles
    c.execute("SELECT COUNT(*) FROM Articles WHERE bin_id=?", (bin_id,))
    nb=c.fetchone()[0]
    if nb==0:
        # plus d'articles => weight=0
        c.execute("UPDATE Pallets SET weight=0 WHERE id=?", (bin_id,))

    conn.commit()
    conn.close()

def edit_article(article_id, new_ref, new_qty, new_login):
    """
    Modifie la reference, la qty, le login.
    Compare new_qty vs old_qty => 
      si new_qty>old_qty => difference = new_qty - old_qty => c'est un 'IN'
      si new_qty<old_qty => difference = old_qty - new_qty => c'est un 'OUT'
    Met à jour Metrics et Movements en conséquence.
    """
    conn=get_db_connection()
    c=conn.cursor()
    c.execute("SELECT bin_id, quantity FROM Articles WHERE id=?", (article_id,))
    row=c.fetchone()
    if not row:
        conn.close()
        return
    bin_id, old_qty=row[0], row[1]

    diff=new_qty - old_qty
    # maj de l'article
    c.execute("""UPDATE Articles
                 SET reference=?, quantity=?, login=?
                 WHERE id=?""",
              (new_ref, new_qty, new_login, article_id))

    now_str=datetime.now().isoformat()
    if diff>0:
        # c'est un IN partiel
        c.execute("UPDATE Metrics SET articles_in=articles_in+? WHERE id=1", (diff,))
        c.execute("""INSERT INTO Movements(article_id, bin_id, action, qty_change, date_time)
                     VALUES(?,?,?,?,?)""",
                  (article_id, bin_id, 'IN', diff, now_str))
    elif diff<0:
        # c'est un OUT partiel
        out_qty = abs(diff)
        c.execute("UPDATE Metrics SET articles_out=articles_out+? WHERE id=1", (out_qty,))
        c.execute("""INSERT INTO Movements(article_id, bin_id, action, qty_change, date_time)
                     VALUES(?,?,?,?,?)""",
                  (article_id, bin_id, 'OUT', out_qty, now_str))

    conn.commit()
    conn.close()

def update_bin_image(bin_id, image_path):
    conn=get_db_connection()
    c=conn.cursor()
    c.execute("UPDATE Pallets SET image_path=? WHERE id=?", (image_path, bin_id))
    conn.commit()
    conn.close()

def remove_bin_image(bin_id):
    conn=get_db_connection()
    c=conn.cursor()
    c.execute("UPDATE Pallets SET image_path=NULL WHERE id=?", (bin_id,))
    conn.commit()
    conn.close()

def search_db(query):
    conn=get_db_connection()
    c=conn.cursor()

    c.execute("SELECT id FROM Pallets WHERE LOWER(bin_name)=?", (query.lower(),))
    row=c.fetchone()
    if row:
        conn.close()
        return ("BIN", query.upper())

    c.execute("""
    SELECT id, bin_id, code, reference, login
    FROM Articles
    WHERE LOWER(code) LIKE ?
    """, (f"%{query.lower()}%",))
    rows=c.fetchall()
    conn.close()
    return ("ARTICLE", rows)

def export_excel_xlsx():
    out_file="export.xlsx"
    conn=get_db_connection()
    c=conn.cursor()

    c.execute("""
    SELECT p.bin_name, p.weight, p.image_path,
           a.id, a.code, a.reference, a.login, a.quantity
    FROM Pallets p
    LEFT JOIN Articles a ON p.id=a.bin_id
    ORDER BY p.bin_name
    """)
    rows=c.fetchall()

    c.execute("SELECT articles_in, articles_out FROM Metrics WHERE id=1")
    metrics=c.fetchone() or (0,0)

    c.execute("SELECT article_id, bin_id, action, qty_change, date_time FROM Movements")
    movements=c.fetchall()
    conn.close()

    wb=openpyxl.Workbook()
    ws=wb.active
    ws.title="Pallets/Articles"
    headers=["BinName","Weight","ImagePath","ArticleID","Code","Reference","Login","Quantity"]
    ws.append(headers)
    for r in rows:
        ws.append(r)

    ws2=wb.create_sheet("Metrics")
    ws2.append(["ArticlesIn","ArticlesOut"])
    ws2.append([metrics[0], metrics[1]])

    ws3=wb.create_sheet("Movements")
    ws3.append(["article_id","bin_id","action","qty_change","date_time"])
    for mv in movements:
        ws3.append(list(mv))

    wb.save(out_file)
    return os.path.abspath(out_file)

def get_group_weight(letter, group_index):
    conn=get_db_connection()
    c=conn.cursor()
    if group_index==1:
        c.execute("SELECT SUM(weight) FROM Pallets WHERE bin_name GLOB ?", (f"{letter}[1-4]",))
    else:
        c.execute("SELECT SUM(weight) FROM Pallets WHERE bin_name GLOB ?", (f"{letter}[5-8]",))
    row=c.fetchone()
    conn.close()
    return row[0] if row and row[0] else 0

def get_total_articles():
    conn=get_db_connection()
    c=conn.cursor()
    c.execute("SELECT COUNT(*) FROM Articles")
    row=c.fetchone()
    conn.close()
    return row[0] if row else 0

def get_metrics():
    conn=get_db_connection()
    c=conn.cursor()
    c.execute("SELECT articles_in, articles_out FROM Metrics WHERE id=1")
    row=c.fetchone()
    conn.close()
    return row if row else (0,0)

def get_movements_in_date_range(start_date, end_date):
    """
    Movements => (id, article_id, bin_id, action, qty_change, date_time)
    """
    conn=get_db_connection()
    c=conn.cursor()
    c.execute("""
    SELECT id, article_id, bin_id, action, qty_change, date_time
    FROM Movements
    WHERE substr(date_time,1,10)>=?
      AND substr(date_time,1,10)<=?
    ORDER BY date_time ASC
    """, (start_date, end_date))
    rows=c.fetchall()
    conn.close()
    return rows

def get_articles_in_multiple_bins():
    conn=get_db_connection()
    c=conn.cursor()
    c.execute("""
    SELECT a.code, p.bin_name
    FROM Articles a
    JOIN Pallets p ON a.bin_id=p.id
    """)
    rows=c.fetchall()
    conn.close()
    d=defaultdict(set)
    for code,bn in rows:
        d[code].add(bn)
    results=[]
    for code,bnset in d.items():
        if len(bnset)>1:
            results.append((code, sorted(list(bnset))))
    return results

def get_top_5_in():
    """
    top 5 articles par la somme de qty_change (action='IN') 
    dans Movements
    """
    conn=get_db_connection()
    c=conn.cursor()
    c.execute("""
    SELECT a.code, SUM(m.qty_change) as total_in
    FROM Movements m
    JOIN Articles a ON m.article_id=a.id
    WHERE m.action='IN'
    GROUP BY m.article_id
    ORDER BY total_in DESC
    LIMIT 5
    """)
    rows=c.fetchall()
    conn.close()
    return rows

def get_top_5_out():
    """
    top 5 articles par la somme de qty_change (action='OUT')
    """
    conn=get_db_connection()
    c=conn.cursor()
    c.execute("""
    SELECT a.code, SUM(m.qty_change) as total_out
    FROM Movements m
    JOIN Articles a ON m.article_id=a.id
    WHERE m.action='OUT'
    GROUP BY m.article_id
    ORDER BY total_out DESC
    LIMIT 5
    """)
    rows=c.fetchall()
    conn.close()
    return rows

def get_movements_by_article_in_range(start_date, end_date):
    """
    total (IN+OUT) par article code, en sommant qty_change
    dans la période
    """
    conn=get_db_connection()
    c=conn.cursor()
    c.execute("""
    SELECT a.code, COUNT(m.id) as total_moves
    FROM Movements m
    JOIN Articles a ON m.article_id=a.id
    WHERE substr(m.date_time,1,10)>=? AND substr(m.date_time,1,10)<=?
    GROUP BY a.code
    ORDER BY total_moves DESC
    """, (start_date, end_date))
    rows=c.fetchall()
    conn.close()
    return rows
